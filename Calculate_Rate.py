import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 文件夹路径
# input_folder = r"D:\新能源预测小组\Project\concat\data\DB+Nas"
input_folder = r"D:\新能源预测小组\Project\concat\data\Fusion"
# output_file = r"D:\新能源预测小组\Project\concat\data\DB和Nas非空数据统计（2021.01.01-今）.xlsx"
output_file = r"D:\新能源预测小组\Project\concat\data\合并数据非空数据统计（2021.01.01-今）.xlsx"

# 创建空的结果字典，用于存储不同Sheet数据
results_sheets = {
                # "Power_DB": pd.DataFrame(), 
                # "Radiation_DB": pd.DataFrame(), 
                # "Power_Nas": pd.DataFrame(), 
                # "Radiation_Nas": pd.DataFrame(),
                "Power_fusion": pd.DataFrame(),
                "Radiation_fusion": pd.DataFrame()
                }

# 定义时间序列的起始时间和间隔
start_time = pd.Timestamp("2021-01-01 00:00")
time_interval = pd.Timedelta(minutes=5)

for file_name in os.listdir(input_folder):
    if file_name.endswith(".csv"):
        file_path = os.path.join(input_folder, file_name)
        
        # 读取文件并解决混合类型警告
        try:
            data = pd.read_csv(file_path, low_memory=False)
        except Exception as e:
            print(f"文件 {file_name} 读取失败: {e}")
            continue

        # 检查是否包含所需列，忽略缺少的文件
        required_columns = [
            # 'Power_DB', 
            # 'Radiation_DB', 
            # 'Power_Nas', 
            # 'Radiation_Nas',
            'Power_fusion',
            'Radiation_fusion'
            ]
        if not set(required_columns).issubset(data.columns):
            print(f"文件 {file_name} 缺少必要列，跳过处理")
            continue

        # 确保时间列为 datetime 格式
        if '时间' not in data.columns:
            data['时间'] = [start_time + i * time_interval for i in range(len(data))]
        else:
            data['时间'] = pd.to_datetime(data['时间'], errors='coerce')  # 转换为 datetime 格式
        
        # 检查时间列是否成功转换
        if data['时间'].isna().all():
            print(f"文件 {file_name} 的 '时间' 列无法解析为有效的日期时间格式，跳过处理")
            continue

        # 提取年月作为分组依据
        data['Time'] = data['时间'].dt.strftime('%Y-%m')

        # 计算每月的非空值率
        monthly_valid_rate = (
            data.groupby(['Time'])[required_columns]
            .apply(lambda x: x.notna().mean().round(3))  # 计算非空值率并保留3位小数
            .reset_index()
        )

        # 添加场站ID列
        station_id = file_name.replace(".csv", "")
        monthly_valid_rate['ID'] = station_id

        # 将数据分别存储到对应的Sheet结构中
        for metric in results_sheets.keys():
            # 提取当前指标数据
            temp_data = monthly_valid_rate.pivot(index='ID', columns='Time', values=metric)
            # 合并到对应的Sheet
            results_sheets[metric] = pd.concat([results_sheets[metric], temp_data], axis=0)

# 将结果写入Excel文件，每个Sheet存储一个指标
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name, sheet_data in results_sheets.items():
        # 转换为百分比格式并保存
        sheet_data = sheet_data.map(lambda x: f"{x * 100:.1f}%" if pd.notna(x) else "")
        sheet_data.to_excel(writer, sheet_name=sheet_name)

# 添加条件格式化（标红 < 80%，标绿 100%）
workbook = load_workbook(output_file)

for sheet_name in results_sheets.keys():
    sheet = workbook[sheet_name]

    # 遍历所有数据单元格，应用条件格式
    for row in range(2, sheet.max_row + 1):  # 从第2行开始（忽略标题）
        for col in range(2, sheet.max_column + 1):  # 从第2列开始（忽略索引）
            cell = sheet.cell(row=row, column=col)
            if cell.value:  # 忽略空值
                value = float(cell.value.strip('%'))  # 去掉百分号并转为浮点数
                if value < 80.0:  # 如果小于80%
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 淡红色
                elif value == 100.0:  # 如果等于100%
                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 淡绿色

# 保存文件
workbook.save(output_file)

print(f"统计完成，结果已保存至 {output_file}，并对80%以下和100%的数据进行了颜色标记")
