import openpyxl
from openpyxl.styles import PatternFill
import os

# 输入文件路径
input_file = r"D:\新能源预测小组\Project\concat\data\a.xlsx"
# 输出文件路径
output_file = r"D:\新能源预测小组\Project\concat\data\b.xlsx"

# 定义颜色填充
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 红色
green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 绿色

# 检查文件路径是否有效
if not os.path.exists(input_file):
    print(f"文件路径无效：{input_file}")
    exit()

try:
    # 加载工作簿
    workbook = openpyxl.load_workbook(input_file)

    # 遍历每个工作表
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"正在处理工作表：{sheet_name}")

        # 遍历工作表内容（假设数据从第2行、第2列开始）
        for row in sheet.iter_rows(min_row=2, min_col=2):  # 数据从第二行第二列开始
            for cell in row:
                if cell.value is None:  # 检查空值
                    cell.value = "0.0%"  # 将空值替换为 "0.0%"
                    cell.fill = red_fill  # 标红
                else:
                    try:
                        # 去掉百分号并转换为浮点数
                        value = float(cell.value.strip('%')) if isinstance(cell.value, str) else cell.value
                        if value < 60.0:  # 如果小于60%
                            cell.fill = red_fill  # 标红
                        elif value >= 90.0:  # 如果大于等于90%
                            cell.fill = green_fill  # 标绿
                    except ValueError:
                        print(f"单元格值无法处理：{cell.value}，跳过")
                        continue

    # 保存处理后的工作簿
    workbook.save(output_file)
    print(f"处理完成，结果保存为：{output_file}")

except Exception as e:
    print(f"文件处理失败: {e}")
